from django.shortcuts import render
import openpyxl
import numpy as np
import pandas as pd
def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        data=pd.read_excel(excel_file)

        # you may put validations here to check extension or file size

        #1wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        #sheets = wb.sheetnames
        #print(sheets)

        # getting a particular sheet
        #worksheet = wb["Sheet1"]
        #print(worksheet)

        # getting active sheet
        #active_sheet = wb.active
        #print(active_sheet)
        info=data.shape
        integer=data.select_dtypes(include=np.number)
        integer=integer.columns.tolist()
        string=data.select_dtypes(exclude=np.number)
        string=string.columns.tolist()

        # reading a cell
        #print(worksheet["A1"].value)

        #excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        #for row in worksheet.iter_rows():
            #row_data = list()
            #for cell in row:
                #row_data.append(str(cell.value))
                #print(cell.value)
            #excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"shape":info,'integer':integer,'string':string})









